# -*- coding: utf-8 -*-
"""Excel 구조 스캔. Claude가 양식을 판단할 재료를 만든다."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def scan_workbook(path: Path, max_rows: int = 60, max_cols: int = 10) -> dict:
    """시트별 셀 값·병합 범위·이미지 수를 훑는다.

    전체를 읽지 않고 앞쪽 일부만 보는 것은, 양식 판단에 필요한 단서(제목 셀,
    헤더 행)가 상단에 몰려 있고 대형 파일에서 리포트가 폭발하는 것을 막기 위해서다.
    """
    path = Path(path)
    wb = load_workbook(path, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        cells = []
        row_limit = min(ws.max_row or 0, max_rows)
        col_limit = min(ws.max_column or 0, max_cols)
        for r in range(1, row_limit + 1):
            for c in range(1, col_limit + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    continue
                text = str(v).strip()
                if not text:
                    continue
                cells.append({"ref": "%s%d" % (get_column_letter(c), r), "value": text})
        sheets.append(
            {
                "name": ws.title,
                "max_row": ws.max_row or 0,
                "max_column": ws.max_column or 0,
                "merged": [str(rng) for rng in ws.merged_cells.ranges],
                "image_count": len(getattr(ws, "_images", [])),
                "cells": cells,
            }
        )
    wb.close()
    return {"file": str(path), "sheets": sheets}
