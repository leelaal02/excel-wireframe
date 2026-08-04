# -*- coding: utf-8 -*-
"""mapping.json에 따라 Excel에서 화면과 상세를 읽는다."""
from __future__ import annotations

import re

from common import Warnings
from openpyxl.utils import column_index_from_string


def _cell_text(ws, row: int, col: int) -> str:
    v = ws.cell(row=row, column=col).value
    return "" if v is None else str(v).strip()


def find_header_row(ws, column: str, marker: str, limit: int = 200) -> int | None:
    """상세 표 헤더 행을 위에서부터 훑어 찾는다.

    헤더 위치는 상단 이미지 크기에 따라 밀리므로 고정할 수 없다.
    """
    col = column_index_from_string(column)
    last = min(ws.max_row or 0, limit)
    for r in range(1, last + 1):
        if _cell_text(ws, r, col) == marker:
            return r
    return None


def parse_screen_meta(text: str, pattern: str | None, fallback_id: str) -> tuple[str, str]:
    """제목 셀에서 화면ID와 화면명을 분리한다.

    패턴이 안 맞으면 셀 전체를 이름으로 두고 ID는 시트명에서 딴다. 양식이 조금
    달라도 화면 자체가 통째로 사라지지 않게 하기 위한 방어다.
    """
    if pattern:
        m = re.search(pattern, text or "")
        if m:
            gd = m.groupdict()
            sid = (gd.get("id") or "").strip()
            name = (gd.get("name") or "").strip()
            if sid or name:
                return sid or fallback_id, name or fallback_id
    return fallback_id, (text or "").strip() or fallback_id


def _read_details(ws, detail_cfg: dict, warns: Warnings, screen_id: str) -> list[dict]:
    marker = detail_cfg.get("header_marker", "No.")
    scan_col = detail_cfg.get("header_scan_column", "A")
    header_row = find_header_row(ws, scan_col, marker)
    if header_row is None:
        warns.add(screen_id, "no-detail",
                  "상세 표 헤더('%s')를 찾지 못했습니다" % marker)
        return []

    cols = {k: column_index_from_string(v) for k, v in detail_cfg["columns"].items()}
    key_col = cols.get("no") or column_index_from_string(scan_col)
    details = []
    r = header_row + 1
    last = ws.max_row or 0
    while r <= last:
        row_vals = {k: _cell_text(ws, r, c) for k, c in cols.items()}
        if not any(row_vals.values()):
            break
        if not row_vals.get("no") and not _cell_text(ws, r, key_col):
            break
        details.append(row_vals)
        r += 1
    return details


def _read_sheet_per_screen(wb, cfg: dict, warns: Warnings) -> list[dict]:
    include = cfg.get("sheet_include")
    rx = re.compile(include) if include else None
    meta_cfg = cfg.get("screen_meta", {})
    meta_cell = meta_cfg.get("cell", "A1")
    pattern = meta_cfg.get("pattern")

    screens = []
    for ws in wb.worksheets:
        if rx and not rx.search(ws.title):
            continue
        raw = ws[meta_cell].value
        text = "" if raw is None else str(raw).strip()
        sid, name = parse_screen_meta(text, pattern, ws.title)
        details = _read_details(ws, cfg["detail"], warns, sid)
        screens.append(
            {
                "id": sid,
                "name": name,
                "sheet": ws.title,
                "images": [],
                "fields": {},
                "details": details,
            }
        )
    return screens


def _read_table(wb, cfg: dict, warns: Warnings) -> list[dict]:
    """1행 = 1화면 양식. 화면ID가 빈 행은 직전 화면의 상세 행으로 귀속시킨다."""
    sheet_name = cfg.get("sheet")
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

    id_col = column_index_from_string(cfg["columns"]["id"])
    name_col = column_index_from_string(cfg["columns"]["name"])
    field_cols = {
        k: column_index_from_string(v) for k, v in (cfg.get("fields") or {}).items()
    }
    detail_cfg = cfg.get("detail") or {"mode": "none"}
    detail_mode = detail_cfg.get("mode", "none")
    detail_cols = {
        k: column_index_from_string(v)
        for k, v in (detail_cfg.get("columns") or {}).items()
    }

    screens: list[dict] = []
    start = int(cfg["header_row"]) + 1
    last = ws.max_row or 0
    blank_streak = 0
    for r in range(start, last + 1):
        sid = _cell_text(ws, r, id_col)
        name = _cell_text(ws, r, name_col)
        detail_vals = {k: _cell_text(ws, r, c) for k, c in detail_cols.items()}
        has_detail = any(detail_vals.values())

        if not sid and not name and not has_detail:
            blank_streak += 1
            if blank_streak >= 3:
                break
            continue
        blank_streak = 0

        if sid:
            screens.append(
                {
                    "id": sid,
                    "name": name or sid,
                    "sheet": ws.title,
                    "images": [],
                    "fields": {
                        k: _cell_text(ws, r, c) for k, c in field_cols.items()
                    },
                    "details": [],
                }
            )
        elif not screens:
            warns.add(None, "orphan-row", "%d행에 화면ID가 없어 건너뜁니다" % r)
            continue

        # merged-cells 양식도 openpyxl로 읽으면 병합 범위의 둘째 행부터 빈 셀이므로
        # grouped-rows와 결과가 같다. 같은 분기로 받는다.
        if detail_mode in ("grouped-rows", "merged-cells") and has_detail:
            screens[-1]["details"].append(detail_vals)

    return screens


def read_screens(wb, mapping: dict, warns: Warnings) -> list[dict]:
    cfg = mapping["excel"]
    layout = cfg.get("layout", "sheet-per-screen")
    if layout == "sheet-per-screen":
        return _read_sheet_per_screen(wb, cfg, warns)
    if layout == "table":
        return _read_table(wb, cfg, warns)
    raise ValueError("지원하지 않는 excel.layout: %s" % layout)
