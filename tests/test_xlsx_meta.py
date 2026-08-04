import datetime
from pathlib import Path

from common import Warnings
from fixtures import make_table_xlsx
from openpyxl import Workbook, load_workbook
from xlsx_meta import find_cover_sheet, read_cover_meta

MAPPING = {"excel": {"layout": "sheet-per-screen", "sheet_include": "^설계_"}}

TABLE_MAPPING = {
    "excel": {
        "layout": "table",
        "sheet": "화면목록",
        "header_row": 3,
        "columns": {"id": "A", "name": "B"},
        "detail": {"mode": "grouped-rows", "columns": {"no": "D", "element": "E", "desc": "F"}},
    }
}


def _cover_xlsx(path: Path, rows=None, title_cells=None) -> Path:
    """실제 샘플과 같은 배치의 표지: 라벨 C열, 값 E열, 단독 셀 B열."""
    wb = Workbook()
    ws = wb.active
    ws.title = "표지"
    for cell, text in (title_cells or {"B3": "화면설계서", "B4": "부제목입니다"}).items():
        ws[cell] = text
    r = 7
    for label, value in (rows or [("프로젝트명", "통합관리시스템"), ("작성일", "2026-06-11")]):
        ws.cell(row=r, column=3, value=label)
        ws.cell(row=r, column=5, value=value)
        r += 1
    wb.create_sheet("설계_SCR001")["A1"] = "화면설계서 - SCR001 (목록)"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def test_find_cover_sheet_skips_screen_sheets(tmp_path: Path):
    wb = load_workbook(_cover_xlsx(tmp_path / "s.xlsx"))
    assert find_cover_sheet(wb, MAPPING) == "표지"


def test_find_cover_sheet_honors_explicit_mapping(tmp_path: Path):
    wb = load_workbook(_cover_xlsx(tmp_path / "s.xlsx"))
    mapping = {"excel": {"sheet_include": "^설계_", "cover": {"sheet": "설계_SCR001"}}}
    assert find_cover_sheet(wb, mapping) == "설계_SCR001"


def test_read_cover_meta_pairs_label_and_value(tmp_path: Path):
    wb = load_workbook(_cover_xlsx(tmp_path / "s.xlsx"))
    meta = read_cover_meta(wb, MAPPING, Warnings())
    assert meta["프로젝트명"] == "통합관리시스템"
    assert meta["작성일"] == "2026-06-11"


def test_read_cover_meta_takes_standalone_cells_as_title_and_subtitle(tmp_path: Path):
    wb = load_workbook(_cover_xlsx(tmp_path / "s.xlsx"))
    meta = read_cover_meta(wb, MAPPING, Warnings())
    assert meta["문서제목"] == "화면설계서"
    assert meta["부제"] == "부제목입니다"


def test_read_cover_meta_ignores_third_standalone_cell(tmp_path: Path):
    xlsx = _cover_xlsx(
        tmp_path / "s.xlsx",
        title_cells={"B3": "제목", "B4": "부제", "B18": "* 주석 문구입니다"},
    )
    meta = read_cover_meta(load_workbook(xlsx), MAPPING, Warnings())
    assert meta["문서제목"] == "제목"
    assert meta["부제"] == "부제"
    assert "* 주석 문구입니다" not in meta.values()


def test_read_cover_meta_normalizes_dates(tmp_path: Path):
    xlsx = _cover_xlsx(tmp_path / "s.xlsx", rows=[("작성일", datetime.date(2026, 6, 11))])
    meta = read_cover_meta(load_workbook(xlsx), MAPPING, Warnings())
    assert meta["작성일"] == "2026-06-11"


def test_meta_overrides_win(tmp_path: Path):
    wb = load_workbook(_cover_xlsx(tmp_path / "s.xlsx"))
    mapping = {"excel": {"sheet_include": "^설계_",
                         "meta_overrides": {"프로젝트명": "손으로 지정한 값"}}}
    meta = read_cover_meta(wb, mapping, Warnings())
    assert meta["프로젝트명"] == "손으로 지정한 값"
    assert meta["작성일"] == "2026-06-11"


def test_read_cover_meta_returns_empty_when_no_cover(tmp_path: Path):
    wb = Workbook()
    wb.active.title = "설계_SCR001"
    wb.active["A1"] = "화면설계서 - SCR001 (목록)"
    p = tmp_path / "only-screens.xlsx"
    wb.save(p)
    loaded = load_workbook(p)
    assert find_cover_sheet(loaded, MAPPING) is None
    assert read_cover_meta(loaded, MAPPING, Warnings()) == {}


def test_read_cover_meta_drops_reserved_labels(tmp_path: Path):
    xlsx = _cover_xlsx(
        tmp_path / "s.xlsx",
        rows=[
            ("프로젝트명", "통합관리시스템"),
            ("source", "짧은 버전.xlsx"),
            ("template", "default.pptx"),
        ],
    )
    meta = read_cover_meta(load_workbook(xlsx), MAPPING, Warnings())
    assert "source" not in meta
    assert "template" not in meta
    assert meta["프로젝트명"] == "통합관리시스템"


def test_meta_overrides_cannot_set_reserved_keys(tmp_path: Path):
    wb = load_workbook(_cover_xlsx(tmp_path / "s.xlsx"))
    mapping = {"excel": {"sheet_include": "^설계_",
                         "meta_overrides": {"source": "hacked.xlsx",
                                            "template": "hacked.pptx"}}}
    meta = read_cover_meta(wb, mapping, Warnings())
    assert "source" not in meta
    assert "template" not in meta
    assert meta["프로젝트명"] == "통합관리시스템"


def test_find_cover_sheet_returns_none_when_named_sheet_missing(tmp_path: Path):
    wb = load_workbook(_cover_xlsx(tmp_path / "s.xlsx"))
    mapping = {"excel": {"sheet_include": "^설계_", "cover": {"sheet": "없는시트"}}}
    assert find_cover_sheet(wb, mapping) is None


def test_read_cover_meta_reads_multiple_pairs_in_one_row(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "표지"
    ws.cell(row=7, column=3, value="프로젝트명")
    ws.cell(row=7, column=5, value="통합관리시스템")
    ws.cell(row=7, column=7, value="상태")
    ws.cell(row=7, column=9, value="진행중")
    wb.create_sheet("설계_SCR001")["A1"] = "화면설계서 - SCR001 (목록)"
    p = tmp_path / "s.xlsx"
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)
    meta = read_cover_meta(load_workbook(p), MAPPING, Warnings())
    assert meta["프로젝트명"] == "통합관리시스템"
    assert meta["상태"] == "진행중"


def test_find_cover_sheet_ignores_non_cover_sheets_when_cover_is_missing(tmp_path: Path):
    """실제 샘플은 표지, 설계_*, 테스트_*, 비교결과요약으로 구성된다. 표지가
    없는 워크북이 들어오면 예전 규칙(sheet_include에 안 걸리는 첫 시트)은
    테스트_*나 비교결과요약을 표지로 오인해 그 시트 텍스트를 meta에 채워
    넣었다 — 경고 하나 없이. 이름 힌트가 없고 라벨-값 쌍도 2개 미만인
    시트는 표지 후보에서 탈락해야 한다."""
    wb = Workbook()
    wb.active.title = "설계_SCR001"
    wb.active["A1"] = "화면설계서 - SCR001 (목록)"
    ws2 = wb.create_sheet("테스트_비고")
    ws2["A1"] = "참고용 메모"
    ws3 = wb.create_sheet("비교결과요약")
    ws3["A1"] = "이전 버전과 비교"
    p = tmp_path / "s.xlsx"
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)
    loaded = load_workbook(p)
    assert find_cover_sheet(loaded, MAPPING) is None
    assert read_cover_meta(loaded, MAPPING, Warnings()) == {}


def test_find_cover_sheet_returns_none_for_table_layout_with_few_pairs(tmp_path: Path):
    """table 레이아웃에서 화면 목록 시트 하나뿐이고 라벨-값 쌍도 2개 미만이면
    (이 자체는 기존 규칙으로도 걸러진다) 표지 없음으로 처리돼야 한다."""
    wb = Workbook()
    ws = wb.active
    ws.title = "화면목록"
    ws["A1"] = "화면 정의서"
    p = tmp_path / "s.xlsx"
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)
    loaded = load_workbook(p)
    assert find_cover_sheet(loaded, TABLE_MAPPING) is None


def test_find_cover_sheet_returns_none_when_table_screen_sheet_has_many_pairs(tmp_path: Path):
    """치명적 결함 1의 회귀 테스트: table 레이아웃은 mapping.excel에
    sheet_include가 없어 예전 코드는 화면 목록 시트(wb.sheetnames[0])를 그대로
    표지 후보로 삼았다. make_table_xlsx의 헤더 행(화면ID/화면명/...)이 라벨-값
    쌍을 트리비얼하게 MIN_COVER_PAIRS 이상으로 채워 채택돼버렸고, 그 결과
    '화면ID' -> '화면명', 'SCR001' -> '이용기관 목록' 같은 화면 데이터가
    문서 meta로 둔갑했다. excel.sheet로 지정된 화면 목록 시트는 쌍이 몇 개든
    후보에서 제외돼야 한다."""
    xlsx = make_table_xlsx(tmp_path / "s.xlsx")
    wb = load_workbook(xlsx)
    assert wb.sheetnames == ["화면목록"]
    assert find_cover_sheet(wb, TABLE_MAPPING) is None
    assert read_cover_meta(wb, TABLE_MAPPING, Warnings()) == {}


def test_find_cover_sheet_skips_hint_named_screen_sheet(tmp_path: Path):
    """중요 발견 3의 회귀 테스트: 이름 힌트 루프(표지/표제/개요/cover/front)는
    sheet_include보다 먼저 도는데, 예전 코드는 그 루프에서 화면 시트 제외를
    전혀 적용하지 않았다. sheet_include='^설계_'인 워크북에서 '설계_개요화면'은
    화면 시트이면서 동시에 '개요' 힌트에 걸려, 실제 표지인 '표지' 시트보다
    먼저(sheetnames 순서상 앞이므로) 잘못 채택됐다."""
    mapping = {"excel": {"layout": "sheet-per-screen", "sheet_include": "^설계_"}}
    wb = Workbook()
    ws = wb.active
    ws.title = "설계_개요화면"
    ws["A1"] = "화면설계서 - SCR001 (개요화면)"
    cover = wb.create_sheet("표지")
    cover["C7"] = "프로젝트명"
    cover["E7"] = "통합관리시스템"
    cover["C8"] = "작성일"
    cover["E8"] = "2026-06-11"
    p = tmp_path / "s.xlsx"
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)
    loaded = load_workbook(p)
    assert loaded.sheetnames[0] == "설계_개요화면"
    assert find_cover_sheet(loaded, mapping) == "표지"


def test_read_cover_meta_ignores_rows_beyond_max_scan_row(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "표지"
    ws.cell(row=7, column=3, value="프로젝트명")
    ws.cell(row=7, column=5, value="통합관리시스템")
    ws.cell(row=41, column=3, value="숨겨진라벨")
    ws.cell(row=41, column=5, value="숨겨진값")
    wb.create_sheet("설계_SCR001")["A1"] = "화면설계서 - SCR001 (목록)"
    p = tmp_path / "s.xlsx"
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)
    meta = read_cover_meta(load_workbook(p), MAPPING, Warnings())
    assert "숨겨진라벨" not in meta
    assert meta["프로젝트명"] == "통합관리시스템"
