from pathlib import Path

from common import read_json, write_json
from extract import main
from fixtures import make_sheet_per_screen_xlsx

MAPPING = {
    "version": 1,
    "excel": {
        "layout": "sheet-per-screen",
        "sheet_include": "^설계_",
        "screen_meta": {
            "cell": "A1",
            "pattern": r"화면설계서\s*-\s*(?P<id>\S+)\s*\((?P<name>.+)\)",
        },
        "detail": {
            "header_scan_column": "A",
            "header_marker": "No.",
            "columns": {"no": "A", "type": "B", "element": "C", "desc": "D", "pos": "E"},
        },
    },
    "template": {"file": "t.pptx", "mode": "clone", "source_slide": 0},
    "options": {"detail_text_source": "desc"},
}

SPEC = [
    {"id": "SCR001", "name": "이용기관 목록", "image": True,
     "details": [{"no": "1", "type": "버튼", "element": "[등록]", "desc": "등록한다", "pos": "상단"}]},
]


def _setup(tmp_path: Path):
    xlsx = make_sheet_per_screen_xlsx(tmp_path / "s.xlsx", SPEC)
    work = tmp_path / "work"
    mp = work / "mapping.json"
    write_json(mp, MAPPING)
    return xlsx, work, mp


def test_extract_creates_screens_json(tmp_path: Path):
    xlsx, work, mp = _setup(tmp_path)
    code = main(["--excel", str(xlsx), "--mapping", str(mp), "--work", str(work)])
    assert code == 0
    data = read_json(work / "screens.json")
    assert data["screens"][0]["id"] == "SCR001"
    assert data["screens"][0]["details"][0]["desc"] == "등록한다"
    assert data["screens"][0]["images"] == ["images/SCR001.png"]
    assert data["meta"]["source"].endswith("s.xlsx")


def test_extract_overwrites_existing_screens_json(tmp_path: Path):
    """screens.json은 중간 산출물이다. 재추출이 그냥 덮어쓴다."""
    xlsx, work, mp = _setup(tmp_path)

    write_json(work / "screens.json", {"meta": {}, "screens": [
        {"id": "OLD", "name": "옛 화면", "images": [], "fields": {}, "details": []}
    ]})

    assert main(["--excel", str(xlsx), "--mapping", str(mp), "--work", str(work)]) == 0

    data = read_json(work / "screens.json")
    assert [s["id"] for s in data["screens"]] == ["SCR001"]
    assert not (work / "screens.new.json").exists()


def test_extract_warns_and_prints_on_zero_screens(tmp_path: Path):
    """sheet_include가 아무 시트에도 안 맞아 화면이 0개 추출되면, 다음 단계인
    build.py의 검증이 잡아주는 것과는 별개로 이 단계에서 바로 원인을 알려줘야
    한다 — 여기가 사람이 sheet_include를 실제로 고칠 수 있는 지점이다."""
    xlsx, work, mp = _setup(tmp_path)
    mapping = read_json(mp)
    mapping["excel"]["sheet_include"] = "^존재하지않는패턴_"
    write_json(mp, mapping)

    import io
    import contextlib

    buf = io.StringIO()
    with contextlib.redirect_stdout(buf):
        code = main(["--excel", str(xlsx), "--mapping", str(mp), "--work", str(work)])
    assert code == 0

    out = buf.getvalue()
    assert "화면 0개" in out
    assert "sheet_include" in out
    data = read_json(work / "screens.json")
    assert data["screens"] == []


def test_extract_puts_cover_meta_into_screens_json(tmp_path: Path):
    xlsx, work, mp = _setup(tmp_path)
    # _setup의 픽스처에는 표지 시트가 '표지' 이름으로 들어 있다
    from openpyxl import load_workbook
    wb = load_workbook(xlsx)
    ws = wb["표지"]
    ws["C7"] = "프로젝트명"
    ws["E7"] = "통합관리시스템"
    ws["C8"] = "작성일"
    ws["E8"] = "2026-06-11"
    wb.save(xlsx)

    assert main(["--excel", str(xlsx), "--mapping", str(mp), "--work", str(work)]) == 0
    meta = read_json(work / "screens.json")["meta"]
    assert meta["프로젝트명"] == "통합관리시스템"
    assert meta["작성일"] == "2026-06-11"
    assert meta["source"].endswith("s.xlsx")


def test_extract_no_longer_hardcodes_title(tmp_path: Path):
    xlsx, work, mp = _setup(tmp_path)
    main(["--excel", str(xlsx), "--mapping", str(mp), "--work", str(work)])
    meta = read_json(work / "screens.json")["meta"]
    assert meta.get("title") != "화면설계서"


def test_extract_reports_when_cover_missing(tmp_path: Path, capsys):
    xlsx, work, mp = _setup(tmp_path)
    from openpyxl import load_workbook
    wb = load_workbook(xlsx)
    del wb["표지"]
    wb.save(xlsx)

    assert main(["--excel", str(xlsx), "--mapping", str(mp), "--work", str(work)]) == 0
    out = capsys.readouterr().out
    assert "표지" in out
    meta = read_json(work / "screens.json")["meta"]
    assert set(meta.keys()) == {"source", "template"}


def test_extract_survives_malformed_cover_config(tmp_path: Path, capsys):
    """치명적 결함 2의 회귀 테스트: mapping.json은 LLM이 스키마 표만 보고
    작성하므로 'cover.sheet'라는 표 항목을 {"cover": "표지"}처럼 평평하게
    잘못 적기 쉽다. 예전 코드는 (cfg.get("cover") or {}).get("sheet")에서
    AttributeError를 그대로 던져 screens.json도, 이미지도, 경고도 없이
    extract.main 전체가 죽었다 — '표지 파싱이 실패해도 화면 페이지 생성은
    정상 진행된다'는 전역 제약을 정면으로 어겼다."""
    xlsx, work, mp = _setup(tmp_path)
    mapping = read_json(mp)
    mapping["excel"]["cover"] = "표지"  # dict가 아니라 문자열 — 잘못된 모양
    write_json(mp, mapping)

    code = main(["--excel", str(xlsx), "--mapping", str(mp), "--work", str(work)])
    assert code == 0

    out = capsys.readouterr().out
    assert "표지" in out
    data = read_json(work / "screens.json")
    assert data["screens"][0]["id"] == "SCR001"
    assert set(data["meta"].keys()) == {"source", "template"}


def test_extract_survives_malformed_meta_overrides(tmp_path: Path, capsys):
    """치명적 결함 2의 두 번째 재현: meta_overrides가 dict가 아니라 list로
    잘못 적히면 overrides.items()가 AttributeError를 던졌다. 이 경우도
    화면 추출은 그대로 완주해야 한다."""
    xlsx, work, mp = _setup(tmp_path)
    mapping = read_json(mp)
    mapping["excel"]["meta_overrides"] = ["프로젝트명"]  # dict가 아니라 list
    write_json(mp, mapping)

    code = main(["--excel", str(xlsx), "--mapping", str(mp), "--work", str(work)])
    assert code == 0

    out = capsys.readouterr().out
    assert "표지" in out
    data = read_json(work / "screens.json")
    assert data["screens"][0]["id"] == "SCR001"
    assert set(data["meta"].keys()) == {"source", "template"}


def test_extract_reserved_keys_always_come_from_arguments(tmp_path: Path, monkeypatch):
    """read_cover_meta는 이미 source/template 라벨을 걸러내므로, 표지에 그
    라벨을 심어봤자 extract.py의 `{**cover_meta, "source": ..., "template":
    ...}` 전개 순서 자체는 한 번도 행사되지 않는다 — 순서를 뒤집어도(즉
    cover_meta가 source/template을 덮어쓰게 바꿔도) 이 방식으로는 검출되지
    않는다. read_cover_meta를 몽키패치해 예약 키가 섞인 dict를 강제로
    돌려주게 만들어, extract.py가 그 값을 실제 인자값으로 덮어쓰는지
    직접 확인한다."""
    xlsx, work, mp = _setup(tmp_path)

    import extract

    def fake_read_cover_meta(wb, mapping, warns):
        return {"source": "가짜값.xlsx", "template": "가짜값.pptx", "프로젝트명": "진짜값"}

    monkeypatch.setattr(extract, "read_cover_meta", fake_read_cover_meta)

    assert main(["--excel", str(xlsx), "--mapping", str(mp), "--work", str(work)]) == 0
    meta = read_json(work / "screens.json")["meta"]
    assert meta["source"].endswith("s.xlsx")
    assert meta["template"] == "t.pptx"
    assert meta["프로젝트명"] == "진짜값"
