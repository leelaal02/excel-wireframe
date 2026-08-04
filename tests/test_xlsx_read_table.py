# -*- coding: utf-8 -*-
from pathlib import Path

from common import Warnings
from fixtures import make_table_xlsx
from openpyxl import load_workbook, Workbook
from xlsx_read import read_screens

MAPPING = {
    "excel": {
        "layout": "table",
        "sheet": "화면목록",
        "header_row": 3,
        "columns": {"id": "A", "name": "B"},
        "fields": {"설명": "C"},
        "detail": {
            "mode": "grouped-rows",
            "columns": {"no": "D", "element": "E", "desc": "F"},
        },
    }
}


def test_table_layout_groups_continuation_rows(tmp_path: Path):
    xlsx = make_table_xlsx(tmp_path / "t.xlsx")
    wb = load_workbook(xlsx, data_only=True)
    screens = read_screens(wb, MAPPING, Warnings())

    assert [s["id"] for s in screens] == ["SCR001", "SCR002"]
    assert screens[0]["name"] == "이용기관 목록"
    assert len(screens[0]["details"]) == 2
    assert screens[0]["details"][1]["element"] == "[삭제] 버튼"
    assert len(screens[1]["details"]) == 1


def test_table_layout_reads_screen_fields(tmp_path: Path):
    xlsx = make_table_xlsx(tmp_path / "t.xlsx")
    wb = load_workbook(xlsx, data_only=True)
    screens = read_screens(wb, MAPPING, Warnings())
    assert screens[0]["fields"] == {"설명": "기관을 조회한다"}


def test_table_layout_without_details(tmp_path: Path):
    xlsx = make_table_xlsx(tmp_path / "t.xlsx")
    wb = load_workbook(xlsx, data_only=True)
    mapping = {"excel": dict(MAPPING["excel"])}
    mapping["excel"]["detail"] = {"mode": "none"}
    screens = read_screens(wb, mapping, Warnings())
    assert screens[0]["details"] == []
    assert len(screens) == 2


def test_table_layout_orphan_row_warning(tmp_path: Path):
    """Orphan detail rows before any screen generate warning and are dropped."""
    # Build workbook with orphan row before first screen
    wb = Workbook()
    ws = wb.active
    ws.title = "화면목록"
    ws["A1"] = "화면 정의서"

    # Header row
    headers = ["화면ID", "화면명", "화면설명", "No.", "요소명", "상세 설명"]
    for i, h in enumerate(headers):
        ws.cell(row=3, column=i + 1, value=h)

    # Row 4: Orphan detail row (no screen ID)
    ws.cell(row=4, column=4, value="1")  # no (detail column)
    ws.cell(row=4, column=5, value="[고아] 버튼")  # element
    ws.cell(row=4, column=6, value="이 행은 어떤 화면에도 속하지 않는다")  # desc

    # Row 5: SCR001
    ws.cell(row=5, column=1, value="SCR001")
    ws.cell(row=5, column=2, value="이용기관 목록")
    ws.cell(row=5, column=3, value="기관을 조회한다")
    ws.cell(row=5, column=4, value="1")
    ws.cell(row=5, column=5, value="[등록] 버튼")
    ws.cell(row=5, column=6, value="등록 팝업을 연다")

    # Row 6: SCR002
    ws.cell(row=6, column=1, value="SCR002")
    ws.cell(row=6, column=2, value="이용기관 상세")
    ws.cell(row=6, column=3, value="상세를 본다")
    ws.cell(row=6, column=4, value="1")
    ws.cell(row=6, column=5, value="[저장] 버튼")
    ws.cell(row=6, column=6, value="변경 내용을 저장한다")

    xlsx = tmp_path / "orphan.xlsx"
    wb.save(str(xlsx))

    # Read with warnings captured
    wb = load_workbook(str(xlsx), data_only=True)
    warns = Warnings()
    screens = read_screens(wb, MAPPING, warns)

    # Assert exactly one orphan-row warning
    warnings_list = warns.to_list()
    orphan_warnings = [w for w in warnings_list if w["code"] == "orphan-row"]
    assert len(orphan_warnings) == 1
    assert orphan_warnings[0]["screen_id"] is None
    assert "4행" in orphan_warnings[0]["message"]

    # Assert orphan row data is not in any screen's details
    all_detail_elements = []
    for screen in screens:
        for detail in screen["details"]:
            all_detail_elements.append(detail.get("element"))
    assert "[고아] 버튼" not in all_detail_elements

    # Assert later screens parse correctly
    assert len(screens) == 2
    assert [s["id"] for s in screens] == ["SCR001", "SCR002"]
    assert screens[0]["details"][0]["element"] == "[등록] 버튼"
