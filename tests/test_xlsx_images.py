import zipfile
from pathlib import Path

import xlsx_images
from common import Warnings
from fixtures import make_png, make_sheet_per_screen_xlsx
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from xlsx_images import (
    _to_png,
    collect_openpyxl_images,
    collect_zip_images,
    extract_images,
)

MAPPING = {"excel": {"layout": "sheet-per-screen"}}

SCREENS_SPEC = [
    {"id": "SCR001", "name": "목록", "image": True,
     "details": [{"no": "1", "type": "버튼", "element": "[등록]", "desc": "등록", "pos": "상단"}]},
    {"id": "SCR002", "name": "상세", "image": True,
     "details": [{"no": "1", "type": "버튼", "element": "[저장]", "desc": "저장", "pos": "하단"}]},
]


def _screens():
    return [
        {"id": "SCR001", "name": "목록", "sheet": "설계_SCR001", "images": [], "fields": {}, "details": []},
        {"id": "SCR002", "name": "상세", "sheet": "설계_SCR002", "images": [], "fields": {}, "details": []},
    ]


def test_collect_openpyxl_images_reads_anchors(tmp_path: Path):
    xlsx = make_sheet_per_screen_xlsx(tmp_path / "s.xlsx", SCREENS_SPEC)
    wb = load_workbook(xlsx)
    found = collect_openpyxl_images(wb)
    assert len(found) == 2
    assert {f["sheet"] for f in found} == {"설계_SCR001", "설계_SCR002"}
    assert found[0]["row"] == 3  # A4 앵커 = 0-based row 3
    assert found[0]["data"][:4] == b"\x89PNG"


def test_collect_zip_images_finds_media(tmp_path: Path):
    xlsx = make_sheet_per_screen_xlsx(tmp_path / "s.xlsx", SCREENS_SPEC)
    found = collect_zip_images(xlsx)
    assert len(found) == 2
    assert all(f["ext"] == "png" for f in found)
    with zipfile.ZipFile(xlsx) as z:
        media = [n for n in z.namelist() if n.startswith("xl/media/")]
    assert len(media) == 2


def test_extract_images_writes_files_and_fills_screens(tmp_path: Path):
    xlsx = make_sheet_per_screen_xlsx(tmp_path / "s.xlsx", SCREENS_SPEC)
    wb = load_workbook(xlsx)
    screens = _screens()
    out = tmp_path / "output"
    extract_images(xlsx, wb, MAPPING, screens, out, Warnings())

    assert screens[0]["images"] == ["images/SCR001.png"]
    assert screens[1]["images"] == ["images/SCR002.png"]
    assert (out / "images" / "SCR001.png").exists()
    assert (out / "images" / "SCR001.png").stat().st_size > 0


def test_extract_images_warns_when_screen_has_none(tmp_path: Path):
    spec = [dict(SCREENS_SPEC[0], image=False)]
    xlsx = make_sheet_per_screen_xlsx(tmp_path / "s.xlsx", spec)
    wb = load_workbook(xlsx)
    screens = [_screens()[0]]
    warns = Warnings()
    extract_images(xlsx, wb, MAPPING, screens, tmp_path / "output", warns)
    assert screens[0]["images"] == []
    assert [w["code"] for w in warns.to_list()] == ["no-image"]


def test_extract_images_falls_back_to_zip_when_openpyxl_misses_images(tmp_path: Path):
    """ws._images를 비워 openpyxl이 찾은 이미지 수 < xl/media 개수인 상황을 실제로 만든다."""
    xlsx = make_sheet_per_screen_xlsx(tmp_path / "s.xlsx", SCREENS_SPEC)
    wb = load_workbook(xlsx)
    for ws in wb.worksheets:
        ws._images = []
    screens = _screens()
    out = tmp_path / "output"
    warns = Warnings()
    extract_images(xlsx, wb, MAPPING, screens, out, warns)

    assert screens[0]["images"] == ["images/SCR001.png"]
    assert screens[1]["images"] == ["images/SCR002.png"]
    assert (out / "images" / "SCR001.png").exists()
    assert (out / "images" / "SCR002.png").exists()
    assert warns.to_list() == []


def test_extract_images_multi_image_screen_gets_suffixed_files(tmp_path: Path):
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("설계_SCR001")
    img1 = make_png(tmp_path / "a.png", color=(255, 0, 0))
    img2 = make_png(tmp_path / "b.png", color=(0, 255, 0))
    ws.add_image(XLImage(str(img1)), "A4")
    ws.add_image(XLImage(str(img2)), "A20")
    xlsx = tmp_path / "multi.xlsx"
    wb.save(xlsx)

    wb2 = load_workbook(xlsx)
    screens = [
        {"id": "SCR001", "name": "목록", "sheet": "설계_SCR001", "images": [], "fields": {}, "details": []}
    ]
    out = tmp_path / "output"
    warns = Warnings()
    extract_images(xlsx, wb2, MAPPING, screens, out, warns)

    assert screens[0]["images"] == ["images/SCR001-1.png", "images/SCR001-2.png"]
    assert (out / "images" / "SCR001-1.png").exists()
    assert (out / "images" / "SCR001-2.png").exists()
    assert warns.to_list() == []


def test_to_png_records_image_convert_failed_and_keeps_original():
    warns = Warnings()
    data, ext = _to_png(b"not an image", "emf", warns, "SCR001")

    assert data == b"not an image"
    assert ext == "emf"
    codes = [w["code"] for w in warns.to_list()]
    assert codes == ["image-convert-failed"]


def test_extract_images_zip_fallback_orders_by_natural_filename(tmp_path: Path):
    """image10.png가 image2.png보다 사전식으로 앞에 오는 함정을 재현한다."""
    ids = ["SCR%02d" % i for i in range(1, 12)]  # 11개 = 두 자릿수 파일명이 섞이는 최소치
    wb = Workbook()
    wb.remove(wb.active)
    originals = {}
    for i, sid in enumerate(ids):
        ws = wb.create_sheet("설계_%s" % sid)
        img_path = tmp_path / ("src_%s.png" % sid)
        color = (10 * i % 256, (255 - 10 * i) % 256, (i * 37) % 256)
        make_png(img_path, color=color)
        originals[sid] = img_path.read_bytes()
        ws.add_image(XLImage(str(img_path)), "A4")
    xlsx = tmp_path / "eleven.xlsx"
    wb.save(xlsx)

    with zipfile.ZipFile(xlsx) as z:
        media = [n for n in z.namelist() if n.startswith("xl/media/")]
    assert len(media) == 11  # 사전식 정렬이면 image10이 image2보다 앞에 와 버그가 드러난다

    wb2 = load_workbook(xlsx)
    for ws in wb2.worksheets:
        ws._images = []  # 모든 시트에서 강제로 zip 폴백을 타게 한다

    screens = [
        {"id": sid, "name": sid, "sheet": "설계_%s" % sid, "images": [], "fields": {}, "details": []}
        for sid in ids
    ]
    out = tmp_path / "output"
    warns = Warnings()
    extract_images(xlsx, wb2, MAPPING, screens, out, warns)

    for sid in ids:
        scr = next(s for s in screens if s["id"] == sid)
        assert scr["images"] == ["images/%s.png" % sid]
        written = (out / "images" / ("%s.png" % sid)).read_bytes()
        assert written == originals[sid]
    assert warns.to_list() == []


def test_collect_openpyxl_images_warns_on_read_failure(tmp_path: Path, monkeypatch):
    xlsx = make_sheet_per_screen_xlsx(tmp_path / "s.xlsx", SCREENS_SPEC)
    wb = load_workbook(xlsx)

    def _boom(img):
        raise RuntimeError("broken ref")

    monkeypatch.setattr(xlsx_images, "_image_bytes", _boom)
    warns = Warnings()
    found = collect_openpyxl_images(wb, warns)

    assert found == []
    codes = [w["code"] for w in warns.to_list()]
    assert codes == ["image-convert-failed", "image-convert-failed"]
