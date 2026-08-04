from pathlib import Path

from common import Warnings
from fixtures import make_sheet_per_screen_xlsx
from openpyxl import Workbook, load_workbook
from xlsx_read import find_header_row, parse_screen_meta, read_screens

MAPPING = {
    "excel": {
        "layout": "sheet-per-screen",
        "sheet_include": "^설계_",
        "screen_meta": {
            "cell": "A1",
            "pattern": r"화면설계서\s*-\s*(?P<id>\S+)\s*\((?P<name>.+)\)",
        },
        "detail": {
            "header_scan_column": "A",
            "header_marker": "No.",
            "columns": {"no": "A", "type": "B", "element": "C", "desc": "D", "pos": "E"},
        },
    }
}

SCREENS = [
    {
        "id": "SCR001",
        "name": "이용기관 목록",
        "image": False,
        "details": [
            {"no": "1", "type": "버튼", "element": "[등록]", "desc": "등록한다", "pos": "상단"},
            {"no": "2", "type": "버튼", "element": "[삭제]", "desc": "삭제한다", "pos": "상단"},
        ],
    },
    {
        "id": "SCR002",
        "name": "이용기관 상세",
        "image": False,
        "details": [
            {"no": "1", "type": "버튼", "element": "[저장]", "desc": "저장한다", "pos": "하단"},
        ],
    },
]


def test_parse_screen_meta_splits_id_and_name():
    pat = MAPPING["excel"]["screen_meta"]["pattern"]
    assert parse_screen_meta("화면설계서 - SCR001 (이용기관 목록)", pat, "설계_SCR001") == (
        "SCR001",
        "이용기관 목록",
    )


def test_parse_screen_meta_falls_back_when_pattern_misses():
    pat = MAPPING["excel"]["screen_meta"]["pattern"]
    assert parse_screen_meta("그냥 제목", pat, "설계_SCR009") == ("설계_SCR009", "그냥 제목")


def test_find_header_row(tmp_path: Path):
    xlsx = make_sheet_per_screen_xlsx(tmp_path / "s.xlsx", SCREENS)
    wb = load_workbook(xlsx, data_only=True)
    ws = wb["설계_SCR001"]
    assert find_header_row(ws, "A", "No.") == 28
    assert find_header_row(ws, "A", "존재하지않음") is None


def test_read_screens_filters_sheets_and_reads_details(tmp_path: Path):
    xlsx = make_sheet_per_screen_xlsx(tmp_path / "s.xlsx", SCREENS)
    wb = load_workbook(xlsx, data_only=True)
    warns = Warnings()
    screens = read_screens(wb, MAPPING, warns)

    assert [s["id"] for s in screens] == ["SCR001", "SCR002"]
    assert screens[0]["name"] == "이용기관 목록"
    assert len(screens[0]["details"]) == 2
    assert screens[0]["details"][1] == {
        "no": "2", "type": "버튼", "element": "[삭제]", "desc": "삭제한다", "pos": "상단",
    }
    assert screens[0]["images"] == []
    assert screens[0]["fields"] == {}


def test_read_details_survives_one_blank_separator_row(tmp_path: Path):
    """상세 표 중간에 빈 구분 행이 한 줄 있어도 표 전체를 읽어야 한다.

    _read_table은 이미 blank_streak(연속 3줄)로 이 문제를 해결했다. _read_details가
    첫 빈 행에서 바로 멈추면, 그 뒤에 남은 상세 행이 통째로 잘려 나간다.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "설계_SCR001"
    ws["A1"] = "화면설계서 - SCR001 (목록)"

    headers = ["No.", "요소타입", "요소명", "상세 설명", "위치"]
    for i, h in enumerate(headers):
        ws.cell(row=10, column=i + 1, value=h)

    rows = [
        ["1", "버튼", "[등록]", "등록한다", "상단"],
        ["2", "버튼", "[삭제]", "삭제한다", "상단"],
        # 표 중간의 빈 구분 행 — 표 끝이 아니다.
        ["", "", "", "", ""],
        ["3", "버튼", "[저장]", "저장한다", "하단"],
    ]
    for j, row in enumerate(rows):
        r = 11 + j
        for i, v in enumerate(row):
            if v:
                ws.cell(row=r, column=i + 1, value=v)

    xlsx = tmp_path / "s.xlsx"
    wb.save(str(xlsx))

    mapping = {
        "excel": {
            "layout": "sheet-per-screen",
            "sheet_include": "^설계_",
            "detail": {
                "header_scan_column": "A",
                "header_marker": "No.",
                "columns": {"no": "A", "type": "B", "element": "C", "desc": "D", "pos": "E"},
            },
        }
    }
    wb2 = load_workbook(xlsx, data_only=True)
    screens = read_screens(wb2, mapping, Warnings())

    assert len(screens) == 1
    details = screens[0]["details"]
    assert [d["no"] for d in details] == ["1", "2", "3"]
    assert details[2]["desc"] == "저장한다"


def test_read_screens_warns_when_header_missing(tmp_path: Path):
    xlsx = make_sheet_per_screen_xlsx(tmp_path / "s.xlsx", SCREENS)
    wb = load_workbook(xlsx, data_only=True)
    mapping = {"excel": dict(MAPPING["excel"])}
    mapping["excel"]["detail"] = dict(MAPPING["excel"]["detail"])
    mapping["excel"]["detail"]["header_marker"] = "없는마커"
    warns = Warnings()
    screens = read_screens(wb, mapping, warns)
    assert screens[0]["details"] == []
    codes = [w["code"] for w in warns.to_list()]
    assert "shape-not-found" not in codes
    assert any(w["message"].startswith("상세 표 헤더") for w in warns.to_list())
